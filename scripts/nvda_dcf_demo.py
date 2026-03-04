"""
NVDA DCF Model — Following Anthropic financial-services-plugins DCF Skill spec
Generates institutional-grade Excel with:
- DCF sheet (projections + sensitivity tables at bottom)
- WACC sheet (CAPM cost of capital)
- Blue fonts for inputs, Black for formulas, Green for cross-sheet links
- Cell comments on all hardcoded inputs
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, numbers
from openpyxl.utils import get_column_letter
from datetime import datetime

wb = openpyxl.Workbook()

# ── Style Definitions ──────────────────────────────────────────────
BLUE = Font(name='Calibri', size=11, color='0000FF')        # Hardcoded inputs
BLUE_BOLD = Font(name='Calibri', size=11, color='0000FF', bold=True)
BLACK = Font(name='Calibri', size=11, color='000000')       # Formulas
BLACK_BOLD = Font(name='Calibri', size=11, color='000000', bold=True)
GREEN = Font(name='Calibri', size=11, color='008000')       # Cross-sheet links
WHITE_BOLD = Font(name='Calibri', size=12, color='FFFFFF', bold=True)
HEADER_FILL = PatternFill('solid', fgColor='1F4E79')        # Dark blue
SUBHEADER_FILL = PatternFill('solid', fgColor='D6E4F0')     # Light blue
INPUT_FILL = PatternFill('solid', fgColor='E2EFDA')         # Light green

THICK = Side(style='medium')
THIN = Side(style='thin')
BORDER_THICK = Border(top=THICK, bottom=THICK, left=THICK, right=THICK)
BORDER_THIN = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)

PCT_FMT = '0.0%'
NUM_FMT = '#,##0'
NUM_FMT_1D = '#,##0.0'
DOLLAR_FMT = '$#,##0.00'

def add_comment(cell, source):
    from openpyxl.comments import Comment
    cell.comment = Comment(source, 'DCF Model')

def style_header_row(ws, row, max_col):
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = WHITE_BOLD
        cell.border = BORDER_THICK

def style_subheader_row(ws, row, max_col):
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = SUBHEADER_FILL
        cell.font = BLACK_BOLD
        cell.border = BORDER_THIN

def write_input(ws, row, col, value, source_text, fmt=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = BLUE
    cell.fill = INPUT_FILL
    cell.border = BORDER_THIN
    add_comment(cell, f'Source: {source_text}')
    if fmt:
        cell.number_format = fmt
    return cell

def write_formula(ws, row, col, value, fmt=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = BLACK
    cell.border = BORDER_THIN
    if fmt:
        cell.number_format = fmt
    return cell

def write_label(ws, row, col, value, bold=False):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = BLACK_BOLD if bold else BLACK
    cell.border = BORDER_THIN
    return cell


# ══════════════════════════════════════════════════════════════════
# SHEET 1: WACC
# ══════════════════════════════════════════════════════════════════
ws_wacc = wb.active
ws_wacc.title = 'WACC'
ws_wacc.column_dimensions['A'].width = 35
ws_wacc.column_dimensions['B'].width = 18

# Title
r = 1
ws_wacc.merge_cells('A1:B1')
c = ws_wacc.cell(row=r, column=1, value='NVIDIA (NVDA) — Cost of Capital (WACC)')
c.fill = HEADER_FILL; c.font = WHITE_BOLD; c.border = BORDER_THICK
ws_wacc.cell(row=r, column=2).fill = HEADER_FILL; ws_wacc.cell(row=r, column=2).border = BORDER_THICK

r = 2
write_label(ws_wacc, r, 1, f'As of {datetime.now().strftime("%Y-%m-%d")}')

# ── CAPM Inputs ──
r = 4
style_subheader_row(ws_wacc, r, 2)
write_label(ws_wacc, r, 1, 'CAPM — Cost of Equity', bold=True)

r = 5; write_label(ws_wacc, r, 1, 'Risk-Free Rate (10Y UST)')
write_input(ws_wacc, r, 2, 0.043, 'FRED DGS10, 2026-02-25', PCT_FMT)

r = 6; write_label(ws_wacc, r, 1, 'Equity Risk Premium')
write_input(ws_wacc, r, 2, 0.055, 'Damodaran ERP, Jan 2026', PCT_FMT)

r = 7; write_label(ws_wacc, r, 1, 'Beta (5Y Monthly)')
write_input(ws_wacc, r, 2, 2.314, 'FMP Profile API, 2026-02-25', NUM_FMT_1D)

r = 8; write_label(ws_wacc, r, 1, 'Cost of Equity (Ke)')
write_formula(ws_wacc, r, 2, '=B5+B7*B6', PCT_FMT)  # 4.3% + 2.314*5.5% = 17.03%

# ── Debt Cost ──
r = 10
style_subheader_row(ws_wacc, r, 2)
write_label(ws_wacc, r, 1, 'Cost of Debt', bold=True)

r = 11; write_label(ws_wacc, r, 1, 'Pre-Tax Cost of Debt')
write_input(ws_wacc, r, 2, 0.035, 'NVDA 10-K FY2025, avg coupon ~3.5%', PCT_FMT)

r = 12; write_label(ws_wacc, r, 1, 'Effective Tax Rate')
write_input(ws_wacc, r, 2, 0.153, 'FMP Income Stmt FY2025: 11.15/72.88+11.15', PCT_FMT)

r = 13; write_label(ws_wacc, r, 1, 'After-Tax Cost of Debt (Kd)')
write_formula(ws_wacc, r, 2, '=B11*(1-B12)', PCT_FMT)

# ── Capital Structure ──
r = 15
style_subheader_row(ws_wacc, r, 2)
write_label(ws_wacc, r, 1, 'Capital Structure', bold=True)

r = 16; write_label(ws_wacc, r, 1, 'Market Cap ($B)')
write_input(ws_wacc, r, 2, 4695.3, 'FMP Quote, 2026-02-25, $B', NUM_FMT_1D)

r = 17; write_label(ws_wacc, r, 1, 'Total Debt ($B)')
write_input(ws_wacc, r, 2, 10.27, 'FMP Balance Sheet FY2025, $B', NUM_FMT_1D)

r = 18; write_label(ws_wacc, r, 1, 'Cash + Short-Term Investments ($B)')
write_input(ws_wacc, r, 2, 43.21, 'FMP Balance Sheet FY2025, $B', NUM_FMT_1D)

r = 19; write_label(ws_wacc, r, 1, 'Net Debt ($B)')
write_formula(ws_wacc, r, 2, '=B17-B18', NUM_FMT_1D)  # Negative = net cash

r = 20; write_label(ws_wacc, r, 1, 'Total Capital (Equity + Debt)')
write_formula(ws_wacc, r, 2, '=B16+B17', NUM_FMT_1D)

r = 21; write_label(ws_wacc, r, 1, 'Equity Weight')
write_formula(ws_wacc, r, 2, '=B16/B20', PCT_FMT)

r = 22; write_label(ws_wacc, r, 1, 'Debt Weight')
write_formula(ws_wacc, r, 2, '=B17/B20', PCT_FMT)

# ── WACC ──
r = 24
style_subheader_row(ws_wacc, r, 2)
write_label(ws_wacc, r, 1, 'Weighted Average Cost of Capital', bold=True)

r = 25; write_label(ws_wacc, r, 1, 'WACC')
write_formula(ws_wacc, r, 2, '=B8*B21+B13*B22', PCT_FMT)

r = 27
write_label(ws_wacc, r, 1, 'Note: NVDA is net cash — WACC ≈ Ke', bold=False)


# ══════════════════════════════════════════════════════════════════
# SHEET 2: DCF
# ══════════════════════════════════════════════════════════════════
ws = wb.create_sheet('DCF', 0)  # Make it the first sheet

# Column widths
ws.column_dimensions['A'].width = 32
for i in range(2, 15):
    ws.column_dimensions[get_column_letter(i)].width = 16

MAX_COL = 13  # A through M

# ── Historical data (FY2021-FY2025) + Projection (FY2026E-FY2030E) ──
# Column layout: A=Label, B=FY2021, C=FY2022, D=FY2023, E=FY2024, F=FY2025, G=FY2026E..K=FY2030E, L=Terminal

HIST_DATA = {
    'revenue':   [16.68, 26.91, 26.97, 60.92, 130.50],
    'cogs':      [6.28,  9.44, 11.62, 16.62,  32.64],
    'gross':     [10.40, 17.48, 15.36, 44.30,  97.86],
    'rd':        [3.92,  5.27,  7.34,  8.68,  12.91],
    'sga':       [1.94,  2.17,  2.44,  2.65,   3.49],
    'oi':        [4.53, 10.04,  4.22, 32.97,  81.45],
    'tax':       [0.08,  0.19, -0.19,  4.06,  11.15],
    'ni':        [4.33,  9.75,  4.37, 29.76,  72.88],
    'da':        [1.10,  1.17,  1.54,  1.51,   1.86],
    'capex':     [1.13,  0.98,  1.83,  1.07,   3.24],
    'sbc':       [1.40,  2.00,  2.71,  3.55,   4.74],
    'opcf':      [5.82,  9.11,  5.64, 28.09,  64.09],
    'fcf':       [4.69,  8.13,  3.81, 27.02,  60.85],
    'ebitda':    [5.69, 11.35,  5.99, 35.58,  86.14],
}

# ── Row positions (pre-planned per skill spec) ──
ROW = {}
ROW['title'] = 1
ROW['subtitle'] = 2
ROW['blank1'] = 3
ROW['scenario_header'] = 4
ROW['scenario_sel'] = 5
ROW['blank2'] = 6

# Scenario assumptions block
ROW['scen_title'] = 7
ROW['scen_years'] = 8
ROW['bear_rev_g'] = 9
ROW['base_rev_g'] = 10
ROW['bull_rev_g'] = 11
ROW['consol_rev_g'] = 12
ROW['blank3'] = 13
ROW['bear_gm'] = 14
ROW['base_gm'] = 15
ROW['bull_gm'] = 16
ROW['consol_gm'] = 17
ROW['blank3b'] = 18
ROW['bear_opex'] = 19
ROW['base_opex'] = 20
ROW['bull_opex'] = 21
ROW['consol_opex'] = 22
ROW['blank4'] = 23

# Historical + Projection
ROW['hist_header'] = 24
ROW['hist_years'] = 25
ROW['revenue'] = 26
ROW['rev_growth'] = 27
ROW['cogs'] = 28
ROW['gross'] = 29
ROW['gm_pct'] = 30
ROW['rd'] = 31
ROW['sga'] = 32
ROW['total_opex'] = 33
ROW['opex_pct'] = 34
ROW['oi'] = 35
ROW['oi_margin'] = 36
ROW['blank5'] = 37

# FCF build
ROW['fcf_header'] = 38
ROW['ebit'] = 39
ROW['tax_on_ebit'] = 40
ROW['nopat'] = 41
ROW['da'] = 42
ROW['capex'] = 43
ROW['nwc'] = 44
ROW['ufcf'] = 45
ROW['blank6'] = 46

# Discounting
ROW['disc_header'] = 47
ROW['wacc_ref'] = 48
ROW['disc_period'] = 49
ROW['disc_factor'] = 50
ROW['pv_fcf'] = 51
ROW['blank7'] = 52

# Terminal Value + Valuation
ROW['tv_header'] = 53
ROW['term_growth'] = 54
ROW['term_fcf'] = 55
ROW['tv'] = 56
ROW['pv_tv'] = 57
ROW['blank8'] = 58

ROW['val_header'] = 59
ROW['sum_pv_fcf'] = 60
ROW['pv_tv_val'] = 61
ROW['ev'] = 62
ROW['net_debt'] = 63
ROW['equity_val'] = 64
ROW['shares'] = 65
ROW['price_per_share'] = 66
ROW['current_price'] = 67
ROW['upside'] = 68
ROW['blank9'] = 69

# Sensitivity tables start at row 71
ROW['sens_header'] = 71

# ── Title Block ──
ws.merge_cells(f'A{ROW["title"]}:K{ROW["title"]}')
c = ws.cell(row=ROW['title'], column=1, value='NVIDIA CORPORATION (NVDA) — DISCOUNTED CASH FLOW ANALYSIS')
c.fill = HEADER_FILL; c.font = WHITE_BOLD; c.border = BORDER_THICK
for col in range(2, 12):
    ws.cell(row=ROW['title'], column=col).fill = HEADER_FILL
    ws.cell(row=ROW['title'], column=col).border = BORDER_THICK

write_label(ws, ROW['subtitle'], 1, f'All figures in $B USD  |  As of {datetime.now().strftime("%Y-%m-%d")}  |  FY ends January')

# ── Scenario Selector ──
style_subheader_row(ws, ROW['scenario_header'], MAX_COL)
write_label(ws, ROW['scenario_header'], 1, 'SCENARIO SELECTION', bold=True)
write_label(ws, ROW['scenario_sel'], 1, 'Active Scenario (1=Bear, 2=Base, 3=Bull)', bold=True)
c = write_input(ws, ROW['scenario_sel'], 2, 2, 'User selection: 1=Bear, 2=Base, 3=Bull')
c.font = BLUE_BOLD

# ── Scenario Assumptions ──
style_subheader_row(ws, ROW['scen_title'], MAX_COL)
write_label(ws, ROW['scen_title'], 1, 'SCENARIO ASSUMPTIONS', bold=True)

# Year headers for projection columns G-K
proj_years = ['FY2026E', 'FY2027E', 'FY2028E', 'FY2029E', 'FY2030E']
for i, yr in enumerate(proj_years):
    write_label(ws, ROW['scen_years'], 7+i, yr, bold=True)

# Revenue growth scenarios
BEAR_G = [0.30, 0.20, 0.12, 0.08, 0.05]
BASE_G = [0.40, 0.28, 0.20, 0.14, 0.10]
BULL_G = [0.50, 0.35, 0.25, 0.18, 0.13]

write_label(ws, ROW['bear_rev_g'], 1, 'Bear — Revenue Growth')
write_label(ws, ROW['base_rev_g'], 1, 'Base — Revenue Growth')
write_label(ws, ROW['bull_rev_g'], 1, 'Bull — Revenue Growth')
write_label(ws, ROW['consol_rev_g'], 1, '→ Active Revenue Growth', bold=True)

for i in range(5):
    col = 7 + i
    write_input(ws, ROW['bear_rev_g'], col, BEAR_G[i], 'Assumption: conservative AI deceleration', PCT_FMT)
    write_input(ws, ROW['base_rev_g'], col, BASE_G[i], 'Assumption: AI growth normalizing', PCT_FMT)
    write_input(ws, ROW['bull_rev_g'], col, BULL_G[i], 'Assumption: AI supercycle sustained', PCT_FMT)
    # Consolidation column: INDEX to pick from scenario
    bear_cell = f'{get_column_letter(col)}{ROW["bear_rev_g"]}'
    base_cell = f'{get_column_letter(col)}{ROW["base_rev_g"]}'
    bull_cell = f'{get_column_letter(col)}{ROW["bull_rev_g"]}'
    write_formula(ws, ROW['consol_rev_g'], col,
                  f'=CHOOSE($B${ROW["scenario_sel"]},{bear_cell},{base_cell},{bull_cell})', PCT_FMT)

# Gross margin scenarios
BEAR_GM = [0.72, 0.70, 0.68, 0.66, 0.65]
BASE_GM = [0.74, 0.73, 0.72, 0.71, 0.70]
BULL_GM = [0.76, 0.76, 0.75, 0.75, 0.74]

write_label(ws, ROW['bear_gm'], 1, 'Bear — Gross Margin')
write_label(ws, ROW['base_gm'], 1, 'Base — Gross Margin')
write_label(ws, ROW['bull_gm'], 1, 'Bull — Gross Margin')
write_label(ws, ROW['consol_gm'], 1, '→ Active Gross Margin', bold=True)

for i in range(5):
    col = 7 + i
    write_input(ws, ROW['bear_gm'], col, BEAR_GM[i], 'Assumption: competition erodes margin', PCT_FMT)
    write_input(ws, ROW['base_gm'], col, BASE_GM[i], 'Assumption: gradual margin normalization', PCT_FMT)
    write_input(ws, ROW['bull_gm'], col, BULL_GM[i], 'Assumption: pricing power sustains', PCT_FMT)
    bear_cell = f'{get_column_letter(col)}{ROW["bear_gm"]}'
    base_cell = f'{get_column_letter(col)}{ROW["base_gm"]}'
    bull_cell = f'{get_column_letter(col)}{ROW["bull_gm"]}'
    write_formula(ws, ROW['consol_gm'], col,
                  f'=CHOOSE($B${ROW["scenario_sel"]},{bear_cell},{base_cell},{bull_cell})', PCT_FMT)

# OpEx % of revenue scenarios
BEAR_OX = [0.14, 0.14, 0.15, 0.15, 0.16]
BASE_OX = [0.12, 0.12, 0.11, 0.11, 0.10]
BULL_OX = [0.10, 0.10, 0.09, 0.09, 0.08]

write_label(ws, ROW['bear_opex'], 1, 'Bear — OpEx % of Revenue')
write_label(ws, ROW['base_opex'], 1, 'Base — OpEx % of Revenue')
write_label(ws, ROW['bull_opex'], 1, 'Bull — OpEx % of Revenue')
write_label(ws, ROW['consol_opex'], 1, '→ Active OpEx % of Revenue', bold=True)

for i in range(5):
    col = 7 + i
    write_input(ws, ROW['bear_opex'], col, BEAR_OX[i], 'Assumption: R&D investment heavy', PCT_FMT)
    write_input(ws, ROW['base_opex'], col, BASE_OX[i], 'Assumption: operating leverage kicks in', PCT_FMT)
    write_input(ws, ROW['bull_opex'], col, BULL_OX[i], 'Assumption: max operating leverage', PCT_FMT)
    bear_cell = f'{get_column_letter(col)}{ROW["bear_opex"]}'
    base_cell = f'{get_column_letter(col)}{ROW["base_opex"]}'
    bull_cell = f'{get_column_letter(col)}{ROW["bull_opex"]}'
    write_formula(ws, ROW['consol_opex'], col,
                  f'=CHOOSE($B${ROW["scenario_sel"]},{bear_cell},{base_cell},{bull_cell})', PCT_FMT)


# ══════════════════════════════════════════════════════════════════
# HISTORICAL + PROJECTION TABLE
# ══════════════════════════════════════════════════════════════════
style_header_row(ws, ROW['hist_header'], MAX_COL)
write_label(ws, ROW['hist_header'], 1, 'INCOME STATEMENT & FREE CASH FLOW', bold=True)
ws.cell(row=ROW['hist_header'], column=1).font = WHITE_BOLD

# Year headers
hist_years = ['FY2021', 'FY2022', 'FY2023', 'FY2024', 'FY2025']
for i, yr in enumerate(hist_years):
    write_label(ws, ROW['hist_years'], 2+i, yr, bold=True)
for i, yr in enumerate(proj_years):
    c = write_label(ws, ROW['hist_years'], 7+i, yr, bold=True)
    c.fill = SUBHEADER_FILL

# ── Revenue ──
write_label(ws, ROW['revenue'], 1, 'Revenue ($B)', bold=True)
for i, v in enumerate(HIST_DATA['revenue']):
    write_input(ws, ROW['revenue'], 2+i, v, f'FMP Income Stmt FY{2021+i}', NUM_FMT_1D)
# Projected revenue
for i in range(5):
    col = 7 + i
    prev_col = get_column_letter(col - 1)
    g_col = get_column_letter(col)
    write_formula(ws, ROW['revenue'], col,
                  f'={prev_col}{ROW["revenue"]}*(1+{g_col}{ROW["consol_rev_g"]})', NUM_FMT_1D)

# Revenue Growth %
write_label(ws, ROW['rev_growth'], 1, '  Growth %')
for i in range(1, 5):
    prev = get_column_letter(2+i-1)
    cur = get_column_letter(2+i)
    write_formula(ws, ROW['rev_growth'], 2+i,
                  f'={cur}{ROW["revenue"]}/{prev}{ROW["revenue"]}-1', PCT_FMT)
for i in range(5):
    col = 7 + i
    write_formula(ws, ROW['rev_growth'], col,
                  f'={get_column_letter(col)}{ROW["consol_rev_g"]}', PCT_FMT)

# COGS
write_label(ws, ROW['cogs'], 1, 'Cost of Revenue ($B)')
for i, v in enumerate(HIST_DATA['cogs']):
    write_input(ws, ROW['cogs'], 2+i, v, f'FMP Income Stmt FY{2021+i}', NUM_FMT_1D)
for i in range(5):
    col = 7 + i
    write_formula(ws, ROW['cogs'], col,
                  f'={get_column_letter(col)}{ROW["revenue"]}*(1-{get_column_letter(col)}{ROW["consol_gm"]})', NUM_FMT_1D)

# Gross Profit
write_label(ws, ROW['gross'], 1, 'Gross Profit ($B)', bold=True)
for i, v in enumerate(HIST_DATA['gross']):
    write_input(ws, ROW['gross'], 2+i, v, f'FMP Income Stmt FY{2021+i}', NUM_FMT_1D)
for i in range(5):
    col = 7 + i
    write_formula(ws, ROW['gross'], col,
                  f'={get_column_letter(col)}{ROW["revenue"]}-{get_column_letter(col)}{ROW["cogs"]}', NUM_FMT_1D)

# Gross Margin %
write_label(ws, ROW['gm_pct'], 1, '  Gross Margin %')
for i in range(5):
    col_l = get_column_letter(2+i)
    write_formula(ws, ROW['gm_pct'], 2+i,
                  f'={col_l}{ROW["gross"]}/{col_l}{ROW["revenue"]}', PCT_FMT)
for i in range(5):
    col = 7 + i
    write_formula(ws, ROW['gm_pct'], col,
                  f'={get_column_letter(col)}{ROW["consol_gm"]}', PCT_FMT)

# R&D
write_label(ws, ROW['rd'], 1, 'R&D Expense ($B)')
for i, v in enumerate(HIST_DATA['rd']):
    write_input(ws, ROW['rd'], 2+i, v, f'FMP Income Stmt FY{2021+i}', NUM_FMT_1D)

# SG&A
write_label(ws, ROW['sga'], 1, 'SG&A Expense ($B)')
for i, v in enumerate(HIST_DATA['sga']):
    write_input(ws, ROW['sga'], 2+i, v, f'FMP Income Stmt FY{2021+i}', NUM_FMT_1D)

# Total OpEx (R&D + SGA) — projected from opex % assumption
write_label(ws, ROW['total_opex'], 1, 'Total Operating Expenses ($B)', bold=True)
for i in range(5):
    col_l = get_column_letter(2+i)
    write_formula(ws, ROW['total_opex'], 2+i,
                  f'={col_l}{ROW["rd"]}+{col_l}{ROW["sga"]}', NUM_FMT_1D)
for i in range(5):
    col = 7 + i
    write_formula(ws, ROW['total_opex'], col,
                  f'={get_column_letter(col)}{ROW["revenue"]}*{get_column_letter(col)}{ROW["consol_opex"]}', NUM_FMT_1D)

# OpEx %
write_label(ws, ROW['opex_pct'], 1, '  OpEx % of Revenue')
for i in range(5):
    col_l = get_column_letter(2+i)
    write_formula(ws, ROW['opex_pct'], 2+i,
                  f'={col_l}{ROW["total_opex"]}/{col_l}{ROW["revenue"]}', PCT_FMT)
for i in range(5):
    col = 7 + i
    write_formula(ws, ROW['opex_pct'], col,
                  f'={get_column_letter(col)}{ROW["consol_opex"]}', PCT_FMT)

# Operating Income
write_label(ws, ROW['oi'], 1, 'Operating Income (EBIT) ($B)', bold=True)
for i, v in enumerate(HIST_DATA['oi']):
    write_input(ws, ROW['oi'], 2+i, v, f'FMP Income Stmt FY{2021+i}', NUM_FMT_1D)
for i in range(5):
    col = 7 + i
    write_formula(ws, ROW['oi'], col,
                  f'={get_column_letter(col)}{ROW["gross"]}-{get_column_letter(col)}{ROW["total_opex"]}', NUM_FMT_1D)

# OI Margin %
write_label(ws, ROW['oi_margin'], 1, '  Operating Margin %')
for i in range(5):
    col_l = get_column_letter(2+i)
    write_formula(ws, ROW['oi_margin'], 2+i,
                  f'={col_l}{ROW["oi"]}/{col_l}{ROW["revenue"]}', PCT_FMT)
for i in range(5):
    col = 7 + i
    write_formula(ws, ROW['oi'], col,
                  f'={get_column_letter(col)}{ROW["gross"]}-{get_column_letter(col)}{ROW["total_opex"]}', NUM_FMT_1D)
    write_formula(ws, ROW['oi_margin'], col,
                  f'={get_column_letter(col)}{ROW["oi"]}/{get_column_letter(col)}{ROW["revenue"]}', PCT_FMT)


# ══════════════════════════════════════════════════════════════════
# FREE CASH FLOW BUILD
# ══════════════════════════════════════════════════════════════════
style_subheader_row(ws, ROW['fcf_header'], MAX_COL)
write_label(ws, ROW['fcf_header'], 1, 'UNLEVERED FREE CASH FLOW', bold=True)

# EBIT (same as OI)
write_label(ws, ROW['ebit'], 1, 'EBIT ($B)')
for i in range(5):
    col = 7 + i
    write_formula(ws, ROW['ebit'], col,
                  f'={get_column_letter(col)}{ROW["oi"]}', NUM_FMT_1D)

# Tax on EBIT
write_label(ws, ROW['tax_on_ebit'], 1, '(-) Tax on EBIT ($B)')
tax_rate_cell = f'WACC!B{12}'  # Tax rate from WACC sheet
for i in range(5):
    col = 7 + i
    c = write_formula(ws, ROW['tax_on_ebit'], col,
                  f'=-{get_column_letter(col)}{ROW["ebit"]}*{tax_rate_cell}', NUM_FMT_1D)
    c.font = GREEN  # Cross-sheet reference

# NOPAT
write_label(ws, ROW['nopat'], 1, 'NOPAT ($B)', bold=True)
for i in range(5):
    col = 7 + i
    write_formula(ws, ROW['nopat'], col,
                  f'={get_column_letter(col)}{ROW["ebit"]}+{get_column_letter(col)}{ROW["tax_on_ebit"]}', NUM_FMT_1D)

# D&A (grow with revenue ~1.5% of rev)
write_label(ws, ROW['da'], 1, '(+) D&A ($B)')
da_pct = write_input(ws, ROW['da'], 1, '(+) D&A ($B)  [~1.4% of rev]', 'FMP CF FY2025: 1.86/130.5=1.4%')
# Overwrite label properly
ws.cell(row=ROW['da'], column=1).value = '(+) D&A ($B)'
for i in range(5):
    col = 7 + i
    write_formula(ws, ROW['da'], col,
                  f'={get_column_letter(col)}{ROW["revenue"]}*0.014', NUM_FMT_1D)
    add_comment(ws.cell(row=ROW['da'], column=col), 'Assumption: D&A ~1.4% of revenue, based on FY2025 ratio')

# CapEx (grow ~2.5% of rev — increasing data center spend)
write_label(ws, ROW['capex'], 1, '(-) CapEx ($B)')
for i in range(5):
    col = 7 + i
    capex_pct = 0.025 + i * 0.003  # 2.5% → 3.7% rising over time
    write_formula(ws, ROW['capex'], col,
                  f'=-{get_column_letter(col)}{ROW["revenue"]}*{capex_pct:.3f}', NUM_FMT_1D)
    add_comment(ws.cell(row=ROW['capex'], column=col), f'Assumption: CapEx {capex_pct*100:.1f}% of rev, rising with scale')

# NWC change (~1% of rev delta)
write_label(ws, ROW['nwc'], 1, '(-) Change in NWC ($B)')
for i in range(5):
    col = 7 + i
    prev_col = get_column_letter(col - 1)
    write_formula(ws, ROW['nwc'], col,
                  f'=-({get_column_letter(col)}{ROW["revenue"]}-{prev_col}{ROW["revenue"]})*0.01', NUM_FMT_1D)
    add_comment(ws.cell(row=ROW['nwc'], column=col), 'Assumption: NWC ~1% of revenue change')

# UFCF
write_label(ws, ROW['ufcf'], 1, 'Unlevered FCF ($B)', bold=True)
for i in range(5):
    col = 7 + i
    cl = get_column_letter(col)
    write_formula(ws, ROW['ufcf'], col,
                  f'={cl}{ROW["nopat"]}+{cl}{ROW["da"]}+{cl}{ROW["capex"]}+{cl}{ROW["nwc"]}', NUM_FMT_1D)


# ══════════════════════════════════════════════════════════════════
# DISCOUNTING
# ══════════════════════════════════════════════════════════════════
style_subheader_row(ws, ROW['disc_header'], MAX_COL)
write_label(ws, ROW['disc_header'], 1, 'DISCOUNTING (Mid-Year Convention)', bold=True)

# WACC reference
write_label(ws, ROW['wacc_ref'], 1, 'WACC (from WACC sheet)')
c = write_formula(ws, ROW['wacc_ref'], 2, f'=WACC!B25', PCT_FMT)
c.font = GREEN

# Discount periods (mid-year: 0.5, 1.5, 2.5, 3.5, 4.5)
write_label(ws, ROW['disc_period'], 1, 'Discount Period')
for i in range(5):
    write_formula(ws, ROW['disc_period'], 7+i, 0.5 + i, NUM_FMT_1D)

# Discount factor = 1/(1+WACC)^period
write_label(ws, ROW['disc_factor'], 1, 'Discount Factor')
for i in range(5):
    col = 7 + i
    write_formula(ws, ROW['disc_factor'], col,
                  f'=1/(1+$B${ROW["wacc_ref"]})^{get_column_letter(col)}{ROW["disc_period"]}', '0.0000')

# PV of FCF
write_label(ws, ROW['pv_fcf'], 1, 'PV of FCF ($B)', bold=True)
for i in range(5):
    col = 7 + i
    cl = get_column_letter(col)
    write_formula(ws, ROW['pv_fcf'], col,
                  f'={cl}{ROW["ufcf"]}*{cl}{ROW["disc_factor"]}', NUM_FMT_1D)


# ══════════════════════════════════════════════════════════════════
# TERMINAL VALUE
# ══════════════════════════════════════════════════════════════════
style_subheader_row(ws, ROW['tv_header'], MAX_COL)
write_label(ws, ROW['tv_header'], 1, 'TERMINAL VALUE (Perpetuity Growth)', bold=True)

write_label(ws, ROW['term_growth'], 1, 'Terminal Growth Rate')
write_input(ws, ROW['term_growth'], 2, 0.035, 'Assumption: long-run AI/semi growth, GDP+ premium', PCT_FMT)

write_label(ws, ROW['term_fcf'], 1, 'Terminal Year FCF ($B)')
write_formula(ws, ROW['term_fcf'], 2, f'=K{ROW["ufcf"]}*(1+B{ROW["term_growth"]})', NUM_FMT_1D)

write_label(ws, ROW['tv'], 1, 'Terminal Value ($B)')
write_formula(ws, ROW['tv'], 2, f'=B{ROW["term_fcf"]}/($B${ROW["wacc_ref"]}-B{ROW["term_growth"]})', NUM_FMT_1D)

write_label(ws, ROW['pv_tv'], 1, 'PV of Terminal Value ($B)')
write_formula(ws, ROW['pv_tv'], 2, f'=B{ROW["tv"]}/(1+$B${ROW["wacc_ref"]})^4.5', NUM_FMT_1D)


# ══════════════════════════════════════════════════════════════════
# EQUITY BRIDGE
# ══════════════════════════════════════════════════════════════════
style_subheader_row(ws, ROW['val_header'], MAX_COL)
write_label(ws, ROW['val_header'], 1, 'ENTERPRISE → EQUITY VALUE', bold=True)

write_label(ws, ROW['sum_pv_fcf'], 1, 'Sum of PV of FCFs ($B)')
write_formula(ws, ROW['sum_pv_fcf'], 2, f'=SUM(G{ROW["pv_fcf"]}:K{ROW["pv_fcf"]})', NUM_FMT_1D)

write_label(ws, ROW['pv_tv_val'], 1, 'PV of Terminal Value ($B)')
write_formula(ws, ROW['pv_tv_val'], 2, f'=B{ROW["pv_tv"]}', NUM_FMT_1D)

write_label(ws, ROW['ev'], 1, 'Enterprise Value ($B)', bold=True)
write_formula(ws, ROW['ev'], 2, f'=B{ROW["sum_pv_fcf"]}+B{ROW["pv_tv_val"]}', NUM_FMT_1D)

write_label(ws, ROW['net_debt'], 1, '(-) Net Debt ($B)  [neg = net cash]')
c = write_input(ws, ROW['net_debt'], 2, -32.94, 'FMP Balance Sheet FY2025: 10.27B debt - 43.21B cash', NUM_FMT_1D)

write_label(ws, ROW['equity_val'], 1, 'Equity Value ($B)', bold=True)
write_formula(ws, ROW['equity_val'], 2, f'=B{ROW["ev"]}-B{ROW["net_debt"]}', NUM_FMT_1D)

write_label(ws, ROW['shares'], 1, 'Diluted Shares Outstanding (B)')
write_input(ws, ROW['shares'], 2, 24.8, 'FMP Income Stmt FY2025, diluted', NUM_FMT_1D)

write_label(ws, ROW['price_per_share'], 1, 'Implied Price Per Share ($)', bold=True)
c = write_formula(ws, ROW['price_per_share'], 2, f'=B{ROW["equity_val"]}/B{ROW["shares"]}', DOLLAR_FMT)
c.font = Font(name='Calibri', size=14, color='000000', bold=True)

write_label(ws, ROW['current_price'], 1, 'Current Market Price ($)')
write_input(ws, ROW['current_price'], 2, 192.85, 'FMP Quote, 2026-02-25', DOLLAR_FMT)

write_label(ws, ROW['upside'], 1, 'Implied Upside / (Downside)', bold=True)
write_formula(ws, ROW['upside'], 2, f'=B{ROW["price_per_share"]}/B{ROW["current_price"]}-1', PCT_FMT)


# ══════════════════════════════════════════════════════════════════
# SENSITIVITY TABLES (3 tables × 5×5 = 75 cells)
# ══════════════════════════════════════════════════════════════════
SENS_START = ROW['sens_header']

# ── Table 1: WACC vs Terminal Growth ──
r = SENS_START
style_header_row(ws, r, 8)
write_label(ws, r, 1, 'SENSITIVITY: WACC vs Terminal Growth Rate', bold=True)
ws.cell(row=r, column=1).font = WHITE_BOLD

r += 1
write_label(ws, r, 1, 'Implied Share Price ($)', bold=True)

# WACC range (columns C-G)
wacc_range = [0.13, 0.15, 0.17, 0.19, 0.21]
# Terminal growth range (rows)
tg_range = [0.02, 0.025, 0.03, 0.035, 0.04]

# Column headers (WACC values)
write_label(ws, r, 2, 'WACC →', bold=True)
for j, w in enumerate(wacc_range):
    c = write_label(ws, r, 3+j, w)
    c.number_format = PCT_FMT
    c.font = BLACK_BOLD

# Row headers (Terminal Growth) + recalc formulas
for i, tg in enumerate(tg_range):
    row = r + 1 + i
    c = write_label(ws, row, 1, 'Terminal Growth')
    c = write_label(ws, row, 2, tg)
    ws.cell(row=row, column=2).number_format = PCT_FMT
    ws.cell(row=row, column=2).font = BLACK_BOLD

    for j, w in enumerate(wacc_range):
        col = 3 + j
        # Full DCF recalc: sum PV FCFs at this WACC + PV TV at this WACC & TG
        # PV FCF = UFCF / (1+wacc)^period
        # TV = terminal_fcf / (wacc - tg), PV_TV = TV / (1+wacc)^4.5
        pv_parts = []
        for k in range(5):
            fcf_cell = f'{get_column_letter(7+k)}{ROW["ufcf"]}'
            period = 0.5 + k
            pv_parts.append(f'{fcf_cell}/(1+{get_column_letter(col)}{r})^{period}')

        term_fcf_ref = f'K{ROW["ufcf"]}*(1+{get_column_letter(2)}{row})'
        tv_formula = f'({term_fcf_ref})/({get_column_letter(col)}{r}-{get_column_letter(2)}{row})'
        pv_tv = f'({tv_formula})/(1+{get_column_letter(col)}{r})^4.5'

        full = f'=({"+".join(pv_parts)}+{pv_tv}-B{ROW["net_debt"]})/B{ROW["shares"]}'
        c = write_formula(ws, row, col, full, DOLLAR_FMT)

        # Highlight base case
        if abs(w - 0.17) < 0.001 and abs(tg - 0.035) < 0.001:
            c.font = Font(name='Calibri', size=11, color='000000', bold=True)
            c.fill = PatternFill('solid', fgColor='FFFF00')

# ── Table 2: Revenue Growth (FY26) vs EBIT Margin (FY26) ──
r2 = r + 1 + len(tg_range) + 2
style_header_row(ws, r2, 8)
write_label(ws, r2, 1, 'SENSITIVITY: FY2026 Revenue Growth vs Operating Margin', bold=True)
ws.cell(row=r2, column=1).font = WHITE_BOLD

r2 += 1
write_label(ws, r2, 1, 'Implied Share Price ($)', bold=True)

rev_g_range = [0.30, 0.35, 0.40, 0.45, 0.50]
opm_range = [0.56, 0.59, 0.62, 0.65, 0.68]

write_label(ws, r2, 2, 'Rev Growth →', bold=True)
for j, rg in enumerate(rev_g_range):
    c = write_label(ws, r2, 3+j, rg)
    c.number_format = PCT_FMT
    c.font = BLACK_BOLD

for i, om in enumerate(opm_range):
    row = r2 + 1 + i
    write_label(ws, row, 1, 'Op Margin')
    c = write_label(ws, row, 2, om)
    ws.cell(row=row, column=2).number_format = PCT_FMT
    ws.cell(row=row, column=2).font = BLACK_BOLD

    for j, rg in enumerate(rev_g_range):
        col = 3 + j
        # Simplified recalc: change FY26 rev and margin, keep FY27-30 from base
        # FY26 Rev = FY25 Rev * (1+rg)
        # FY26 EBIT = FY26 Rev * om
        # FY26 UFCF = EBIT*(1-tax) + D&A - CapEx - NWC
        # Then FY27-30 use base growth from there
        fy25_rev = 130.5
        rg_ref = f'{get_column_letter(col)}{r2}'
        om_ref = f'{get_column_letter(2)}{row}'

        # Build chain: FY26 from variable, FY27-30 from base assumptions
        # This is a simplified recalc for the sensitivity
        base_g = [0.28, 0.20, 0.14, 0.10]
        base_gm = [0.73, 0.72, 0.71, 0.70]
        base_ox = [0.12, 0.11, 0.11, 0.10]
        tax_r = 0.153

        # Build the full formula as a nested calculation
        # For readability, compute FY26 UFCF from inputs, then use base for FY27-30
        # FY26: rev=130.5*(1+rg), ebit=rev*om, ufcf=ebit*(1-0.153)+rev*0.014-rev*0.025-drev*0.01
        formula_parts = []
        # FY26
        formula_parts.append(
            f'(({fy25_rev}*(1+{rg_ref}))*{om_ref}*(1-{tax_r})'
            f'+{fy25_rev}*(1+{rg_ref})*0.014'
            f'-{fy25_rev}*(1+{rg_ref})*0.025'
            f'-{fy25_rev}*{rg_ref}*0.01)'
            f'/(1+$B${ROW["wacc_ref"]})^0.5'
        )
        # FY27-30: chain from FY26 using base assumptions
        prev_rev = f'{fy25_rev}*(1+{rg_ref})'
        for k in range(4):
            this_rev = f'{prev_rev}*(1+{base_g[k]})'
            gm = base_gm[k]
            ox = base_ox[k]
            opm_k = gm - ox
            period = 1.5 + k
            ufcf = f'({this_rev}*{opm_k}*(1-{tax_r})+{this_rev}*0.014-{this_rev}*{0.025+k*0.003:.3f}-({this_rev}-{prev_rev})*0.01)'
            formula_parts.append(f'{ufcf}/(1+$B${ROW["wacc_ref"]})^{period}')
            prev_rev = this_rev

        # Terminal value
        last_rev = prev_rev
        tv = f'({last_rev}*{base_gm[3]-base_ox[3]}*(1-{tax_r})*(1+$B${ROW["term_growth"]}))/($B${ROW["wacc_ref"]}-$B${ROW["term_growth"]})'
        pv_tv = f'({tv})/(1+$B${ROW["wacc_ref"]})^4.5'

        full = f'=({"+".join(formula_parts)}+{pv_tv}-B{ROW["net_debt"]})/B{ROW["shares"]}'

        c = write_formula(ws, row, col, full, DOLLAR_FMT)
        # Highlight base case
        if abs(rg - 0.40) < 0.001 and abs(om - 0.62) < 0.001:
            c.font = Font(name='Calibri', size=11, color='000000', bold=True)
            c.fill = PatternFill('solid', fgColor='FFFF00')


# ── Table 3: Beta vs Risk-Free Rate ──
r3 = r2 + 1 + len(opm_range) + 2
style_header_row(ws, r3, 8)
write_label(ws, r3, 1, 'SENSITIVITY: Beta vs Risk-Free Rate', bold=True)
ws.cell(row=r3, column=1).font = WHITE_BOLD

r3 += 1
write_label(ws, r3, 1, 'Implied Share Price ($)', bold=True)

beta_range = [1.5, 1.8, 2.1, 2.3, 2.6]
rfr_range = [0.035, 0.038, 0.043, 0.048, 0.053]

write_label(ws, r3, 2, 'Beta →', bold=True)
for j, b in enumerate(beta_range):
    c = write_label(ws, r3, 3+j, b)
    c.number_format = NUM_FMT_1D
    c.font = BLACK_BOLD

for i, rf in enumerate(rfr_range):
    row = r3 + 1 + i
    write_label(ws, row, 1, 'Risk-Free Rate')
    c = write_label(ws, row, 2, rf)
    ws.cell(row=row, column=2).number_format = PCT_FMT
    ws.cell(row=row, column=2).font = BLACK_BOLD

    for j, b in enumerate(beta_range):
        col = 3 + j
        beta_ref = f'{get_column_letter(col)}{r3}'
        rfr_ref = f'{get_column_letter(2)}{row}'
        erp = 0.055  # ERP fixed

        # wacc_calc = rfr + beta * erp (simplified, ~100% equity)
        # Then full DCF recalc with this WACC
        pv_parts = []
        for k in range(5):
            fcf_cell = f'{get_column_letter(7+k)}{ROW["ufcf"]}'
            period = 0.5 + k
            pv_parts.append(f'{fcf_cell}/(1+{rfr_ref}+{beta_ref}*{erp})^{period}')

        term_fcf_ref = f'K{ROW["ufcf"]}*(1+$B${ROW["term_growth"]})'
        tv_formula = f'({term_fcf_ref})/({rfr_ref}+{beta_ref}*{erp}-$B${ROW["term_growth"]})'
        pv_tv = f'({tv_formula})/(1+{rfr_ref}+{beta_ref}*{erp})^4.5'

        full = f'=({"+".join(pv_parts)}+{pv_tv}-B{ROW["net_debt"]})/B{ROW["shares"]}'
        c = write_formula(ws, row, col, full, DOLLAR_FMT)

        # Highlight base case
        if abs(b - 2.3) < 0.05 and abs(rf - 0.043) < 0.001:
            c.font = Font(name='Calibri', size=11, color='000000', bold=True)
            c.fill = PatternFill('solid', fgColor='FFFF00')


# ── Freeze panes and print settings ──
ws.freeze_panes = f'A{ROW["hist_years"]+1}'
ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)

# ── Save ──
output_path = '/Users/owen/CC workspace/Finance/reports/NVDA_DCF_Model_2026-02-25.xlsx'
wb.save(output_path)
print(f'✅ DCF Model saved to: {output_path}')
print(f'   Sheets: {wb.sheetnames}')
print(f'   DCF rows: {r3 + len(rfr_range) + 1}')
