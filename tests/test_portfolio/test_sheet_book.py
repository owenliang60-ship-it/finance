"""Tests for portfolio.holdings.sheet_book — synthetic fixtures only.

PRIVACY: repo is public. Never put real tickers' real amounts or the real
sheet ID in this file. All data below is fictional.
"""
import datetime as dt
import io

import openpyxl
import pytest

from portfolio.holdings.sheet_book import (
    SheetBookError,
    parse_sheet_book,
)

HEADER = [
    "Market", "Stock Ticker", "Google Price", "Manual Price", "Manual Price $",
    "Last Price", "Last DPS", "Yield on Cost", "Last Price Yield", "Shares",
    "Cost", "Cost (Per Share)", "Unrealized Gain/Loss", "Unrealized Gain/Loss (%)",
    "Realized Gain/Loss", "Momentums Collected", "Total Gain/Loss", "Mkt Value",
    "Category",
]

FETCHED_AT = dt.datetime(2026, 6, 10, 12, 0, 0)


def _row(market, ticker, last_price, shares, cost_ps, mkt_value, category):
    r = [None] * len(HEADER)
    r[0], r[1], r[5], r[9], r[11], r[17], r[18] = (
        market, ticker, last_price, shares, cost_ps, mkt_value, category)
    return r


DEFAULT_ROWS = [
    _row("US", "AAA", 10.0, 100, 8.0, 1000.0, "Sentiment"),
    _row("KR", "000001", 50.0, 20, 55.0, 1000.0, "Fundamental"),
    _row("HK", "HKG:0001", 5.0, 200, 4.5, 1000.0, "Value"),
    _row("US", "BBB LEAPS", 30.0, 10, 25.0, 300.0, "LTH"),
    _row("US", "BSKT", 20.0, 50, 15.0, 1000.0, "Fundamental"),
    _row("KR", "000002", 12.0, 0, 0.0, 0.0, "Momentum"),   # closed -> filtered
    _row("Crypto", "", 0.0, 0, 0.0, 0.0, ""),               # template row -> filtered
]


def build_fixture(rows=None, cash_label="现金合计", cash=111000.0,
                  include_sheet26=True, total_label="total", total=1000000.0,
                  include_summary=True, noise_rows=1):
    wb = openpyxl.Workbook()
    ws = wb.active
    if include_summary:
        ws.title = "Summary_OSV"
        for _ in range(noise_rows):
            ws.append(["junk note row"])
        ws.append(HEADER)
        for r in (rows if rows is not None else DEFAULT_ROWS):
            ws.append(r)
    else:
        ws.title = "Other"
    ps = wb.create_sheet("Portfolio Summary")
    ps.append(["持仓市值", 999.0])
    ps.append([cash_label, cash])
    if include_sheet26:
        s26 = wb.create_sheet("Sheet26")
        s26.append([None, "美股", 500000.0, 0.5])
        s26.append([None, total_label, total, 1])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


class TestParseHoldings:
    def test_parses_and_filters(self):
        book = parse_sheet_book(build_fixture(), FETCHED_AT)
        syms = {h.symbol for h in book.holdings}
        assert syms == {"AAA", "000001", "0001", "BBB LEAPS", "BSKT"}
        assert book.fetched_at == FETCHED_AT

    def test_hkg_prefix_stripped_and_fields(self):
        book = parse_sheet_book(build_fixture(), FETCHED_AT)
        hk = {h.symbol: h for h in book.holdings}["0001"]
        assert hk.raw_ticker == "HKG:0001"
        assert hk.market == "HK"
        assert hk.shares == 200
        assert hk.cost_per_share == 4.5
        assert hk.sheet_price == 5.0
        assert hk.category == "Value"

    def test_leaps_flag(self):
        book = parse_sheet_book(build_fixture(), FETCHED_AT)
        flags = {h.symbol: h.is_leaps for h in book.holdings}
        assert flags["BBB LEAPS"] is True
        assert flags["AAA"] is False

    def test_markets_and_sheet_prices_sidecars(self):
        book = parse_sheet_book(build_fixture(), FETCHED_AT)
        assert book.markets()["000001"] == "KR"
        assert book.sheet_prices()["BSKT"] == 20.0

    def test_duplicate_ticker_merged_weighted(self):
        rows = [
            _row("US", "AAA", 10.0, 100, 8.0, 1000.0, "Sentiment"),
            _row("US", "AAA", 10.0, 100, 12.0, 1000.0, "Sentiment"),
        ]
        book = parse_sheet_book(build_fixture(rows=rows), FETCHED_AT)
        assert len(book.holdings) == 1
        h = book.holdings[0]
        assert h.shares == 200
        assert h.cost_per_share == pytest.approx(10.0)
        assert h.market_value == 2000.0

    def test_empty_holdings_raises(self):
        rows = [_row("KR", "000002", 12.0, 0, 0.0, 0.0, "Momentum")]
        with pytest.raises(SheetBookError, match="no holdings"):
            parse_sheet_book(build_fixture(rows=rows), FETCHED_AT)

    def test_missing_summary_tab_raises(self):
        with pytest.raises(SheetBookError, match="Summary_OSV"):
            parse_sheet_book(build_fixture(include_summary=False), FETCHED_AT)


class TestCashAndCapital:
    def test_cash_parsed(self):
        book = parse_sheet_book(build_fixture(), FETCHED_AT)
        assert book.cash_usd == 111000.0

    def test_missing_cash_label_raises(self):
        with pytest.raises(SheetBookError, match="现金合计"):
            parse_sheet_book(build_fixture(cash_label="别的"), FETCHED_AT)

    def test_total_capital_from_sheet26(self):
        book = parse_sheet_book(build_fixture(), FETCHED_AT)
        assert book.total_capital_usd == 1000000.0

    def test_missing_total_returns_none(self):
        book = parse_sheet_book(build_fixture(total_label="not-total"), FETCHED_AT)
        assert book.total_capital_usd is None

    def test_missing_sheet26_tab_returns_none(self):
        book = parse_sheet_book(build_fixture(include_sheet26=False), FETCHED_AT)
        assert book.total_capital_usd is None


class TestLoadSheetBook:
    def test_missing_env_raises(self, monkeypatch):
        from portfolio.holdings.sheet_book import load_sheet_book
        monkeypatch.delenv("PORTFOLIO_SHEET_ID", raising=False)
        with pytest.raises(SheetBookError, match="not configured"):
            load_sheet_book()

    def test_http_error_message_has_no_sheet_id(self, monkeypatch):
        from portfolio.holdings.sheet_book import load_sheet_book
        fake_id = "FAKE_SHEET_ID_123"
        monkeypatch.setenv("PORTFOLIO_SHEET_ID", fake_id)

        class FakeResp:
            status_code = 404
            content = b""

        monkeypatch.setattr("requests.get", lambda url, timeout: FakeResp())
        with pytest.raises(SheetBookError) as ei:
            load_sheet_book()
        assert fake_id not in str(ei.value)

    def test_network_error_message_has_no_url(self, monkeypatch):
        from portfolio.holdings.sheet_book import load_sheet_book
        fake_id = "FAKE_SHEET_ID_123"
        monkeypatch.setenv("PORTFOLIO_SHEET_ID", fake_id)

        def boom(url, timeout):
            raise ConnectionError("https://docs.google.com/x/" + fake_id)

        monkeypatch.setattr("requests.get", boom)
        with pytest.raises(SheetBookError) as ei:
            load_sheet_book()
        assert fake_id not in str(ei.value)
        assert "docs.google.com" not in str(ei.value)

    def test_network_error_chain_suppressed(self, monkeypatch):
        from portfolio.holdings.sheet_book import load_sheet_book
        fake_id = "FAKE_SHEET_ID_123"
        monkeypatch.setenv("PORTFOLIO_SHEET_ID", fake_id)

        def boom(url, timeout):
            raise ConnectionError("https://docs.google.com/x/" + fake_id)

        monkeypatch.setattr("requests.get", boom)
        with pytest.raises(SheetBookError) as ei:
            load_sheet_book()
        assert ei.value.__suppress_context__ is True
        assert ei.value.__cause__ is None

    def test_success_path(self, monkeypatch):
        from portfolio.holdings.sheet_book import load_sheet_book
        monkeypatch.setenv("PORTFOLIO_SHEET_ID", "FAKE")
        payload = build_fixture()

        class FakeResp:
            status_code = 200
            content = payload

        monkeypatch.setattr("requests.get", lambda url, timeout: FakeResp())
        book = load_sheet_book()
        assert book.cash_usd == 111000.0
        assert {h.symbol for h in book.holdings} >= {"AAA", "BSKT"}
