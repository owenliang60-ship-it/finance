"""Google Sheet book-of-record reader (read-only, in-memory).

PRIVACY: the sheet ID is a bearer secret (this repo is public).
Never log or embed the sheet ID / URL in exception messages.

P3 purity: this module never writes to company.db / market.db.
"""
from __future__ import annotations

import datetime as dt
import io
import logging
import os
from dataclasses import dataclass
from typing import Dict, List, Optional

logger = logging.getLogger(__name__)

SUMMARY_TAB = "Summary_OSV"
CASH_TAB = "Portfolio Summary"
CAPITAL_TAB = "Sheet26"
CASH_LABEL = "现金合计"
CAPITAL_LABEL = "total"
MAX_HEADER_SCAN_ROWS = 15
ALLOWED_MARKETS = {"US", "HK", "KR", "CRYPTO"}
REQUIRED_COLUMNS = [
    "Market", "Stock Ticker", "Last Price", "Shares",
    "Cost (Per Share)", "Mkt Value", "Category",
]


class SheetBookError(Exception):
    """Sheet download/parse failure. Message must never contain sheet ID/URL."""


@dataclass
class SheetHolding:
    symbol: str            # normalized: "HKG:0700" -> "0700", upper/strip
    raw_ticker: str        # sheet original (debug only — never logged)
    market: str            # "US" / "KR" / "HK" / "CRYPTO"
    shares: float
    cost_per_share: float  # sheet "Cost (Per Share)"
    sheet_price: float     # sheet "Last Price" (already USD)
    market_value: float    # sheet "Mkt Value" (USD, informational)
    category: str          # sheet "Category" sleeve
    is_leaps: bool


@dataclass
class SheetBook:
    holdings: List[SheetHolding]
    cash_usd: float
    total_capital_usd: Optional[float]
    fetched_at: dt.datetime

    def markets(self) -> Dict[str, str]:
        return {h.symbol: h.market for h in self.holdings}

    def sheet_prices(self) -> Dict[str, float]:
        return {h.symbol: h.sheet_price for h in self.holdings}


def _to_float(value) -> float:
    if value is None:
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def _to_float_strict(value, context: str) -> float:
    """Parse a critical numeric field; raise instead of silently coercing.

    Book-of-record fields must never degrade to 0 on formula errors
    (e.g. '#N/A') or export anomalies.
    """
    if value is None or (isinstance(value, str) and value.strip() == ""):
        raise SheetBookError("missing %s" % context)
    try:
        return float(value)
    except (TypeError, ValueError):
        raise SheetBookError("non-numeric %s" % context) from None


def _normalize_symbol(raw: str) -> str:
    sym = str(raw).strip().upper()
    if sym.startswith("HKG:"):
        sym = sym[4:]
    return sym


def _parse_holdings(ws) -> List[SheetHolding]:
    header_map = None
    data_rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if header_map is None:
            if i >= MAX_HEADER_SCAN_ROWS:
                break
            cells = [str(c).strip() if c is not None else "" for c in row]
            if "Stock Ticker" in cells:
                header_map = {name: idx for idx, name in enumerate(cells) if name}
            continue
        data_rows.append(row)
    if header_map is None:
        raise SheetBookError(
            "header row with 'Stock Ticker' not found in %s" % SUMMARY_TAB)

    missing = [c for c in REQUIRED_COLUMNS if c not in header_map]
    if missing:
        raise SheetBookError(
            "missing columns in %s: %s" % (SUMMARY_TAB, ", ".join(missing)))

    def cell(row, name):
        idx = header_map[name]
        return row[idx] if idx < len(row) else None

    merged: Dict[str, SheetHolding] = {}
    for row in data_rows:
        raw_ticker = cell(row, "Stock Ticker")
        if raw_ticker is None or str(raw_ticker).strip() == "":
            continue
        raw = str(raw_ticker).strip()
        sym = _normalize_symbol(raw)
        shares = _to_float_strict(
            cell(row, "Shares"), "Shares for %s in %s" % (raw, SUMMARY_TAB))
        if shares <= 0:
            continue  # closed position (realized-only row, no price checks)
        market = str(cell(row, "Market") or "").strip().upper()
        if not market:
            raise SheetBookError("missing Market for %s in %s" % (raw, SUMMARY_TAB))
        if market not in ALLOWED_MARKETS:
            raise SheetBookError(
                "unknown Market '%s' for %s in %s" % (market, raw, SUMMARY_TAB))
        h = SheetHolding(
            symbol=sym,
            raw_ticker=raw,
            market=market,
            shares=shares,
            cost_per_share=_to_float_strict(
                cell(row, "Cost (Per Share)"),
                "Cost (Per Share) for %s" % raw),
            sheet_price=_to_float_strict(
                cell(row, "Last Price"), "Last Price for %s" % raw),
            market_value=_to_float(cell(row, "Mkt Value")),
            category=str(cell(row, "Category") or "").strip(),
            is_leaps=sym.endswith(" LEAPS"),
        )
        if sym in merged:
            prev = merged[sym]
            total_shares = prev.shares + h.shares
            if total_shares > 0:
                prev.cost_per_share = (
                    prev.cost_per_share * prev.shares
                    + h.cost_per_share * h.shares
                ) / total_shares
            prev.shares = total_shares
            prev.market_value += h.market_value
            logger.warning("duplicate ticker merged: %s", sym)
            # sheet_price/category: keep first row's value (first-wins,
            # warning logged above)
        else:
            merged[sym] = h
    return list(merged.values())


def _parse_cash(ws) -> float:
    for row in ws.iter_rows(values_only=True):
        label = str(row[0]).strip() if row and row[0] is not None else ""
        if label == CASH_LABEL:
            value = _to_float_strict(
                row[1] if len(row) > 1 else None,
                "cash value in %s" % CASH_TAB)
            if value < 0:
                raise SheetBookError("cash value invalid in %s" % CASH_TAB)
            return value
    raise SheetBookError("label not found in %s: %s" % (CASH_TAB, CASH_LABEL))


def _parse_total_capital(ws) -> Optional[float]:
    """Total capital from Sheet26. None only when the tab/label is absent
    (PI falls back to env); a present-but-broken value must raise."""
    for row in ws.iter_rows(values_only=True):
        if row is None or len(row) < 2 or row[1] is None:
            continue
        if str(row[1]).strip().lower() == CAPITAL_LABEL:
            value = _to_float_strict(
                row[2] if len(row) > 2 else None,
                "total value in %s" % CAPITAL_TAB)
            if value <= 0:
                raise SheetBookError("total value invalid in %s" % CAPITAL_TAB)
            return value
    return None


def parse_sheet_book(xlsx_bytes: bytes, fetched_at: dt.datetime) -> SheetBook:
    import openpyxl

    try:
        wb = openpyxl.load_workbook(
            io.BytesIO(xlsx_bytes), data_only=True, read_only=True)
    except Exception as e:
        # 200 + HTML (login/quota page) arrives here as BadZipFile etc.
        raise SheetBookError(
            "sheet content is not a valid xlsx: %s" % type(e).__name__) from None
    try:
        if SUMMARY_TAB not in wb.sheetnames:
            raise SheetBookError("sheet tab missing: %s" % SUMMARY_TAB)
        if CASH_TAB not in wb.sheetnames:
            raise SheetBookError("sheet tab missing: %s" % CASH_TAB)

        holdings = _parse_holdings(wb[SUMMARY_TAB])
        if not holdings:
            raise SheetBookError("no holdings parsed from sheet")
        cash = _parse_cash(wb[CASH_TAB])
        total = (
            _parse_total_capital(wb[CAPITAL_TAB])
            if CAPITAL_TAB in wb.sheetnames else None
        )
    finally:
        wb.close()
    return SheetBook(
        holdings=holdings, cash_usd=cash,
        total_capital_usd=total, fetched_at=fetched_at,
    )


def load_sheet_book(sheet_id: Optional[str] = None,
                    timeout: float = 30.0) -> SheetBook:
    """Download the book-of-record workbook and parse it.

    Raises SheetBookError on any failure. Error messages never contain
    the sheet ID or URL (bearer secret, public repo).
    """
    import requests

    sheet_id = (sheet_id or os.environ.get("PORTFOLIO_SHEET_ID", "")).strip()
    if not sheet_id:
        raise SheetBookError("PORTFOLIO_SHEET_ID not configured")
    url = (
        "https://docs.google.com/spreadsheets/d/%s/export?format=xlsx"
        % sheet_id
    )
    try:
        resp = requests.get(url, timeout=timeout)
    except Exception as e:
        # requests exception messages may embed the URL — keep class name only
        raise SheetBookError(
            "sheet download failed: %s" % type(e).__name__) from None
    if resp.status_code != 200:
        raise SheetBookError(
            "sheet download failed: HTTP %d" % resp.status_code)
    return parse_sheet_book(resp.content, dt.datetime.now(dt.timezone.utc))
